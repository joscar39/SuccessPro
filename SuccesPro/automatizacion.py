import unittest
import time
import openpyxl
import openpyxl as xl
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By


class Automatizacion(unittest.TestCase):

    def setUp(self):
        self.driver = webdriver.Chrome(executable_path=r"C:\Users\user\Documents\Proyectos\Pruebas"
                                                       r"\SuccesPro\recursos\driver\chromedriver.exe")

    def test_TC_001_registrousuario(self):
        # inicializando manejador web
        driver = self.driver
        driver.get("https://www.automationexercise.com/")
        driver.maximize_window()
        # validando que se muestre el modulo de login / signup
        driver.find_element(By.XPATH, "/html/body/header/div/div/div/div[2]/div/ul/li[4]/a").click()
        time.sleep(1)
        self.assertEqual("Automation Exercise - Signup / Login", driver.title, "No Se logro ingresar "
                                                                               "al modulo de registro de Usuario")
        # inicializando load_workbook para usar los datos excel
        filesheet = "C:/Users/user/Documents/Proyectos/Pruebas/SuccesPro/recursos/datos.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        # Guardando en variables los datos del archivo excel
        i = 0
        username, email, gender, password = datos[f'A{i}:D{i}'][0]
        day, month, year, sub_notice = datos[f'E{i}:H{i}'][0]
        rece_offer, name, lastname = datos[f'I{i}:K{i}'][0]
        company, addre, addre2, country = datos[f'L{i}:O{i}'][0]
        state, city, codezip, mnumber = datos[f'P{i}:S{i}'][0]
        # ingresar datos de username y correo para avanzar al formulario de registro
        driver.find_element(By.NAME, "name").send_keys(username.value)
        driver.find_element(By.XPATH, "//*[@id='form']/div/div/div[3]/div/form/input[3]").send_keys(email.value)
        driver.get_screenshot_as_file("C:\\Users\\user\\Documents\\Proyectos\\Pruebas"
                                      "\\SuccesPro\\recursos\\screenshots\\TC-001_acceder_form_registro.png")
        driver.find_element(By.XPATH, "/html/body/section/div/div/div[3]/div/form/button").click()
        time.sleep(1)
        self.assertEqual("Automation Exercise - Signup", driver.title, "No se logro ingresar "
                                                                       "al modulo de registro de Usuario, datos "
                                                                       "incorrectos o usuarios ya se "
                                                                       "encuentra registrado")
        # llenado de informacion complementaria del formulario
        if gender.value == "mr":
            driver.find_element(By.ID, "id_gender1").click()
        else:
            driver.find_element(By.ID, "id_gender2").click()
        driver.find_element(By.ID, "password").send_keys(password.value)
        time.sleep(1)
        selectday = driver.find_element(By.ID, "days")
        opcion = selectday.find_elements(By.TAG_NAME, "option")
        time.sleep(0.25)
        for i in opcion:
            i.click()
            time.sleep(0.00001)
        seleccionar = Select(selectday)
        seleccionar.select_by_value(str(day.value))
        time.sleep(1)
        selectmonth = driver.find_element(By.ID, "months")
        opcion2 = selectmonth.find_elements(By.TAG_NAME, "option")
        time.sleep(0.25)
        for i in opcion2:
            i.click()
            time.sleep(0.00001)
        seleccionar2 = Select(selectmonth)
        seleccionar2.select_by_value(str(month.value))
        time.sleep(1)
        selectyear = driver.find_element(By.ID, "years")
        opcion3 = selectyear.find_elements(By.TAG_NAME, "option")
        time.sleep(0.25)
        for i in opcion3:
            i.click()
        time.sleep(0.00001)
        seleccionar3 = Select(selectyear)
        seleccionar3.select_by_value(str(year.value))
        time.sleep(1)
        driver.get_screenshot_as_file("C:\\Users\\user\\Documents\\Proyectos\\"
                                      "Pruebas\\SuccesPro\\recursos\\screenshots\\TC-001_formulario_1era_parte.png")
        driver.execute_script("window.scrollTo(0, 300);")

        if sub_notice.value == "yes":
            newsletter = driver.find_element(By.ID, "newsletter")
            newsletter.click()
            confirm = newsletter.is_selected()
            if confirm is True:
                print("Se ha marcado la opcion para recibir noticias")
            else:
                print("Error el check no se ha seleccionado aun")
        else:
            print("El usuario no desea recibir noticias del sitio web")

        if rece_offer.value == "yes":
            offert = driver.find_element(By.ID, "optin")
            offert.click()
            confir2 = offert.is_selected()

            if confir2 is True:
                print("La opcion recibir ofertas ha sido marcada correctamente ")
            else:
                print("Error no se ha marcado la opcion de recibir ofertas")
        else:
            print("El usuario no desea recibir notificaciones de ofertas")

        driver.execute_script("window.scrollTo(0, 600);")

        driver.find_element(By.ID, "first_name").send_keys(name.value)
        driver.find_element(By.ID, "last_name").send_keys(lastname.value)
        driver.find_element(By.ID, "company").send_keys(str(company.value))
        driver.find_element(By.ID, "address1").send_keys(str(addre.value))
        driver.find_element(By.ID, "address2").send_keys(str(addre2.value))

        cctry = driver.find_element(By.ID, "country")
        opc = cctry.find_elements(By.TAG_NAME, "option")
        time.sleep(0.25)
        for i in opc:
            i.click()
            time.sleep(0.00001)
        sele = Select(cctry)
        sele.select_by_value(country.value)
        driver.find_element(By.ID, "state").send_keys(state.value)
        driver.find_element(By.ID, "city").send_keys(city.value)
        driver.find_element(By.ID, "zipcode").send_keys(str(codezip.value))
        driver.find_element(By.ID, "mobile_number").send_keys(str(mnumber.value))
        time.sleep(1)
        driver.get_screenshot_as_file("C:\\Users\\user\\Documents\\"
                                      "Proyectos\\Pruebas\\SuccesPro\\recursos\\screenshots\\"
                                      "TC-001_formulario_2da_parte.png")
        driver.find_element(By.XPATH, "/html/body/section/div/div/div/div/form/button").click()
        # Verificacion de registro exitoso
        account = driver.find_element(By.XPATH, "//b[contains(text(), 'Account Created')]")
        self.assertTrue(account, "Registro Exitoso")

        driver.get_screenshot_as_file("C:\\Users\\user\\Documents\\Proyectos\\Pruebas\\"
                                      "SuccesPro\\recursos\\screenshots\\TC-001_Registro_exitoso.png")

        driver.find_element(By.CSS_SELECTOR, "#form > div > div > div > div > a").click()

        driver.get_screenshot_as_file("C:\\Users\\user\\Documents\\Proyectos\\Pruebas\\SuccesPro"
                                      "\\recursos\\screenshots\\TC-001_Usuariologgeado.png")
        wb.close()

    def test_TC_002_iniciar_usuario_ya_existente(self):

        # llamando el manejador web
        driver = self.driver
        driver.get("https://www.automationexercise.com/")
        driver.maximize_window()
        driver.get("https://www.automationexercise.com/")
        # validando que se muestre el modulo de login / signup
        driver.find_element(By.XPATH, "/html/body/header/div/div/div/div[2]/div/ul/li[4]/a").click()
        time.sleep(1)
        self.assertEqual("Automation Exercise - Signup / Login", driver.title, "No Se logro ingresar "
                                                                               "al modulo de registro de Usuario")
        # inicializando load_workbook para usar los datos excel
        filesheet = "C:/Users/user/Documents/Proyectos/Pruebas/SuccesPro/recursos/datos.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        # Guardando en variables los datos del archivo excel
        i = 0
        username, email = datos[f'A{i}:B{i}'][0]

        driver.find_element(By.NAME, "name").send_keys(username.value)
        driver.find_element(By.XPATH, "//*[@id='form']/div/div/div[3]/div/form/input[3]").send_keys(email.value)
        driver.find_element(By.XPATH, "/html/body/section/div/div/div[3]/div/form/button").click()
        time.sleep(1)

        existe = driver.find_element(By.XPATH, "//p[contains(text(), 'Email Address already exist')]")
        self.assertTrue(existe, "El Correo ingresado ya se encuentra en uso")

        driver.get_screenshot_as_file("C:\\Users\\user\\Documents\\Proyectos\\Pruebas"
                                      "\\SuccesPro\\recursos\\screenshots\\TC-002_Usuario_ya_esta_registrado.png")
        wb.close()

    def test_TC_003_iniciar_sesion(self):
        # llamando el manejador web
        driver = self.driver
        driver.get("https://www.automationexercise.com/")
        driver.maximize_window()
        driver.get("https://www.automationexercise.com/")
        # validando que se muestre el modulo de login / signup
        driver.find_element(By.XPATH, "/html/body/header/div/div/div/div[2]/div/ul/li[4]/a").click()
        time.sleep(1)
        self.assertEqual("Automation Exercise - Signup / Login", driver.title, "No Se logro ingresar "
                                                                               "al modulo de registro de Usuario")
        # inicializando load_workbook para usar los datos excel
        filesheet = "C:/Users/user/Documents/Proyectos/Pruebas/SuccesPro/recursos/datos.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        # Guardando en variables los datos del archivo excel
        mail = datos['B1']
        password = datos['D1']
        driver.find_element(By.XPATH, "/html/body/section/div/div/div[1]/div/form/input[2]").send_keys(mail.value)
        driver.find_element(By.NAME, "password").send_keys(password.value)
        # Se utilizara un xpath para el identificador del boton login ya que no se puede
        # seleccionar por class name u otro atributo
        driver.find_element(By.XPATH, "/html/body/section/div/div/div[1]/div/form/button").click()
        time.sleep(1)
        # Buscar texto indicativo de usuario loggeado
        login = driver.find_element(By.XPATH, "//a[contains(text(), 'Logged in as')]")
        # Verificacion sobre el texto loggeado esta disponible lo que regresa un true y termina el proceso
        self.assertTrue(login, "Inicio de sesion Exitoso ")
        driver.get_screenshot_as_file("C:\\Users\\user\\Documents\\Proyectos\\Pruebas"
                                      "\\SuccesPro\\recursos\\screenshots\\TC-003_Inicio_de_sesion.png")
        # Eliminar cuenta
        driver.find_element(By.XPATH, "//a[contains(text(), ' Delete Account')]").click()
        eliminar = driver.find_element(By.XPATH, "//h1[contains(text(), 'Delete Account')]")
        self.assertTrue(eliminar, "No se pulso la opcion eliminar")
        driver.get_screenshot_as_file("C:\\Users\\user\\Documents\\Proyectos\\Pruebas"
                                      "\\SuccesPro\\recursos\\screenshots\\TC-003_Eliminar_cuenta.png")
        wb.close()

    def test_TC_004_Buscar_productos_y_ver_detalle(self):
        # llamando el manejador web
        driver = self.driver
        driver.get("https://www.automationexercise.com/")
        driver.maximize_window()
        driver.get("https://www.automationexercise.com/")
        # validando que se muestre el modulo de login / signup
        driver.find_element(By.XPATH, "/html/body/header/div/div/div/div[2]/div/ul/li[4]/a").click()
        time.sleep(1)
        self.assertEqual("Automation Exercise - Signup / Login", driver.title, "No Se logro ingresar "
                                                                               "al modulo de registro de Usuario")
        # inicializando load_workbook para usar los datos excel
        filesheet = "C:/Users/user/Documents/Proyectos/Pruebas/SuccesPro/recursos/datos.xlsx"
        wb = openpyxl.load_workbook(filesheet)
        datos = wb["Hoja1"]
        # Guardando en variables los datos del archivo excel

        mail = datos['B1']
        password = datos['D1']
        driver.find_element(By.XPATH, "/html/body/section/div/div/div[1]/div/form/input[2]").send_keys(mail.value)
        driver.find_element(By.NAME, "password").send_keys(password.value)
        # Se utilizara un xpath para el identificador del boton login ya que no se puede
        # seleccionar por class name u otro atributo
        driver.find_element(By.XPATH, "/html/body/section/div/div/div[1]/div/form/button").click()
        time.sleep(1)

        # Acceder a listado de productos
        driver.find_element(By.XPATH, "//a[contains(text(), 'Products')]").click()

        listado = driver.find_element(By.XPATH, "//h2[contains(text(), 'All Products')]").is_displayed()

        if listado is True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

            driver.get_screenshot_as_file("C:\\Users\\user\\Documents\\Proyectos\\Pruebas"
                                          "\\SuccesPro\\recursos\\screenshots\\TC-004_Ultimo_Producto_de_la_lista.png")
            driver.find_element(By.XPATH, "/html/body/section[2]/div/div/div[2]/div[1]/div[35]"
                                          "/div/div[2]/ul/li/a").click()
            detalle = driver.find_element(By.XPATH, "//h2[contains(text(), 'GRAPHIC "
                                                    "DESIGN MEN T SHIRT - BLUE')]")
            driver.get_screenshot_as_file("C:\\Users\\user\\Documents\\Proyectos\\Pruebas"
                                          "\\SuccesPro\\recursos\\screenshots\\TC-004_Detalle_del_producto.png")
            self.assertTrue(detalle, "No se visualiza el detalle del producto")
        else:
            self.assertEqual("All Products",
                             driver.find_element(By.XPATH, "//h2[contains(text(),"
                                                           " 'All Products')]"), "No se encontro"
                                                                                 " el listado de productos")

        time.sleep(1)
        wb.close()
        # Eliminar el dato del usuario registrado del excel
        wb = xl.load_workbook(filesheet)
        ws = wb.active
        ws.delete_rows(1)  # para la fila 1
        wb.save(filesheet)

    def tearDown(self):
        self.driver.close()


if __name__ == '__main__':
    unittest.main()
